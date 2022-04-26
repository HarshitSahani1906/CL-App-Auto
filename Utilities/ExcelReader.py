import openpyxl
from Excel_Files.ExcelPath import GetExcelpath
from Utilities.generatingLogs import Logging


class ExcelOps:
    global obj_GetExcel
    obj_GetExcel = GetExcelpath()
    global logger
    logger=Logging().log()

    def getRowCount(path, sheetname):
        workbook = openpyxl.load_workbook(obj_GetExcel.testcasesheet())
        sheet = workbook["Login"]
        logger.info("Number of Excel rows retrieved")
        return sheet.max_row

    def getColCount(path, sheetname):
        workbook = openpyxl.load_workbook(obj_GetExcel.testcasesheet())
        sheet = workbook["Login"]
        logger.info("Number of Excel column retrieved")
        return sheet.max_column

    def getCellData(path, sheetname, rownum, colnum):
        workbook = openpyxl.load_workbook(obj_GetExcel.testcasesheet())
        sheet = workbook["Login"]
        logger.info(" Excel cell value retrieved")
        return sheet.cell(row=rownum, column=colnum).value

    def setCellData(path, sheetname, rownum, colnum, data):
        workbook = openpyxl.load_workbook(obj_GetExcel.testcasesheet())
        sheet = workbook["Login"]
        logger.info("Excel cell value written")
        sheet.cell(row=rownum, column=colnum).value = data
        workbook.save(path)




